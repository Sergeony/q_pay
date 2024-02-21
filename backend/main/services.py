import logging
from decimal import Decimal
from typing import List

import openpyxl
from asgiref.sync import async_to_sync
from channels.db import database_sync_to_async
from channels.layers import get_channel_layer
from django.db.transaction import atomic
from django.utils import timezone
from pandas import DataFrame

from .exceptions import InsufficientBalanceError
from .models import (
    User,
    Transaction,
    Balance,
    PrevTransactionTraders,
    BalanceHistory,
)
from .serializers import (
    TransactionSerializer,
    BalanceSerializer,
    APITransactionSerializer,
    ClientTransactionStatusUpdateSerializer,
)
from api.tasks import process_transaction_task


logger = logging.getLogger(__name__)


def create_transactions_excel(transactions: List[Transaction], transaction_type='input'):
    # TODO: modify in according to the current transaction model
    """
    Prepare an exel file with transactions report.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Transactions'

    if transaction_type == 'input':
        columns = ['Transaction ID', 'Status', 'Merchant', 'Bank', 'Requisites', 'Claimed Amount', 'Actual Amount',
                   'Date']
    elif transaction_type == 'output':
        columns = ['Transaction ID', 'Status', 'Amount', 'Bank', 'Card Number', 'Receipt URL', 'Date']
    else:
        raise ValueError("Invalid transaction type. Must be 'input' or 'output'.")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, transaction in enumerate(transactions, 2):
        worksheet.cell(row=row_num, column=1).value = str(transaction.id)
        worksheet.cell(row=row_num, column=2).value = transaction.get_status_display()

        if transaction_type == 'input':
            worksheet.cell(row=row_num, column=3).value = transaction.merchant.id
            worksheet.cell(row=row_num, column=4).value = transaction.trader_bank_details.bank.title
            worksheet.cell(row=row_num, column=5).value = transaction.trader_bank_details.card_number
            worksheet.cell(row=row_num, column=6).value = transaction.amount
            worksheet.cell(row=row_num, column=7).value = transaction.actual_amount or 0
        elif transaction_type == 'output':
            worksheet.cell(row=row_num, column=3).value = transaction.amount
            if transaction.trader_bank_details:
                value = transaction.trader_bank_details.bank.title
            else:
                value = 'N/A'
            worksheet.cell(row=row_num, column=4).value = value
            worksheet.cell(row=row_num, column=5).value = transaction.client_card_number

        worksheet.cell(row=row_num, column=8).value = transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')

    return workbook


def get_eligible_trader_ids_for_transactions(transactions: List[Transaction]) -> List[str]:
    # TODO: modify in according to the changes
    """
    Get traders eligible to handle the transactions.

    Search filters:
        1. The trader is active
        2. The trader has active ads with the banks of the transactions
        3. The every single ad has active bank details in itself with eligible unused limits for the transactions
        # TODO: implement
    """
    eligible_trader_ids = User.objects.filter(
        type=User.Type.TRADER,
        is_active=True
    ).values_list('id', flat=True)

    for transaction in transactions:
        eligible_trader_ids = eligible_trader_ids.filter(
            advertisements__is_active=True,
            advertisements__bank_id=transaction.trader_bank_details.bank,
        ).values_list('id', flat=True)

        if not eligible_trader_ids:
            break

    return list(eligible_trader_ids)


def notify_user_on_transaction_update(transaction: Transaction, is_admin=False):
    """
    Send an event about a new transaction to notify the trader and provide the transaction data.
    """
    transaction_data = TransactionSerializer(transaction).data

    if is_admin:
        group_name = 'admin'
    else:
        group_name = f'user_{transaction.trader.id}'

    channel_layer = get_channel_layer()
    message = {
        'type': 'send_transaction_to_user',
        'transaction_data': transaction_data,
    }
    async_to_sync(channel_layer.group_send)(group_name, message)  # TODO: check fixes to avoid warning


def notify_user_on_balance_update(balance: Balance):
    balance_data = BalanceSerializer(balance).data

    if balance.user is None:
        group_name = 'admin'
    else:
        group_name = f'user_{balance.user.id}'

    channel_layer = get_channel_layer()
    message = {
        'type': 'send_balance_to_user',
        'balance_data': balance_data,
    }
    async_to_sync(channel_layer.group_send)(group_name, message)


def notify_client_on_transaction_update(transaction: Transaction):
    transaction_data = APITransactionSerializer(transaction).data

    channel_layer = get_channel_layer()
    message = {
        'type': 'send_transaction_to_client',
        'transaction_data': transaction_data,
    }
    group_name = f'user_{transaction.merchant.id}'
    async_to_sync(channel_layer.group_send)(group_name, message)  # TODO: check fixes to avoid warning


def freeze_user_balance_for_transaction(user_balance: Balance, transaction: Transaction):
    """
    Freeze the stated transaction amount on the user's balance until the end of the transaction.
    """
    new_active_balance = user_balance.active_balance - transaction.amount
    new_frozen_balance = user_balance.frozen_balance + transaction.amount
    if new_active_balance < Decimal('0.00'):
        raise InsufficientBalanceError("Insufficient balance to perform the transaction.")

    user_balance.active_balance = new_active_balance
    user_balance.frozen_balance = new_frozen_balance
    user_balance.save()

    BalanceHistory.objects.create(
        user=user_balance.user,
        new_active_balance=user_balance.active_balance,
        new_frozen_balance=user_balance.frozen_balance,
        change_active_balance_amount=-transaction.amount,
        change_frozen_balance_amount=transaction.amount,
        change_reason=BalanceHistory.ChangeReason.FREEZE_FOR_TRANSACTION
    )


def update_balances_on_successful_transaction(sender_balance: Balance, transaction: Transaction):
    """
    Perform balance transfers between the users if the transaction ends successfully.

    Return the penalty amount to the sender's balance if the transaction actual amount is less than stated amount.

    Freeze the amount to refund on the sender's balance if the transaction actual amount exceeds the stated amount.

    Calculate the amount to be transferred equal the actual amount
        if it does not exceed the stated amount,
        otherwise of the stated amount,
        without taking into account the commission amounts.

    Pay to the service and to the trader the amounts of the
        commissions of the amount to be transferred.

    Calculate the net amount according to the transaction type:
        1. Subtract the service commission amount from the amount to be transferred.
        2. Subtract the trader commission amount as well, if the trader is sender (for deposit transactions).

    Remove the stated amount from the sender's balance and transfer the net amount to the receiver's balance.
    """
    amount_to_transfer = min(transaction.actual_amount, transaction.amount)
    trader_commission_amount = transaction.trader_commission * amount_to_transfer / 100
    service_commission_amount = transaction.service_commission * amount_to_transfer / 100
    total_commission_amount = transaction.amount + service_commission_amount

    if transaction.status == Transaction.Status.PARTIAL:
        penalty_amount = transaction.amount - transaction.actual_amount
        sender_balance.active_balance += penalty_amount
        sender_balance.save()
        BalanceHistory.objects.create(
            user=sender_balance.user,
            new_active_balance=sender_balance.active_balance,
            new_frozen_balance=sender_balance.frozen_balance,
            change_active_balance_amount=penalty_amount,
            change_reason=BalanceHistory.ChangeReason.RETURN_PENALTY
        )
    elif transaction.status == Transaction.Status.REFUND_REQUESTED:
        amount_to_refund = transaction.actual_amount - transaction.amount
        sender_balance.frozen_balance += amount_to_refund
        sender_balance.save()
        BalanceHistory.objects.create(
            user=sender_balance.user,
            new_active_balance=sender_balance.active_balance,
            new_frozen_balance=sender_balance.frozen_balance,
            change_frozen_balance_amount=amount_to_refund,
            change_reason=BalanceHistory.ChangeReason.FREEZE_FOR_REFUND
        )
    elif transaction.status == Transaction.Status.COMPLETED:
        transaction.actual_amount = transaction.amount

    bank_details = transaction.trader_bank_details
    if transaction.type == transaction.Type.DEPOSIT:
        bank_details.current_daily_turnover += transaction.actual_amount
        bank_details.current_weekly_turnover += transaction.actual_amount
        bank_details.current_monthly_turnover += transaction.actual_amount
    elif transaction.type == transaction.Type.WITHDRAWAL:
        bank_details.current_daily_turnover += (transaction.actual_amount - total_commission_amount)  # TODO:consider it
        bank_details.current_weekly_turnover += (transaction.actual_amount - total_commission_amount)
        bank_details.current_monthly_turnover += (transaction.actual_amount - total_commission_amount)
    bank_details.save()

    service_balance = Balance.objects.select_for_update().get(user=None)
    service_balance.active_balance += service_commission_amount
    service_balance.save()
    BalanceHistory.objects.create(
        user=service_balance.user,
        new_active_balance=service_balance.active_balance,
        new_frozen_balance=service_balance.frozen_balance,
        change_active_balance_amount=service_commission_amount,
        change_reason=BalanceHistory.ChangeReason.PAY_COMMISSION
    )

    net_amount = amount_to_transfer - service_commission_amount
    if transaction.type == Transaction.Type.DEPOSIT:
        sender_balance.active_balance += trader_commission_amount
        sender_balance.save()
        BalanceHistory.objects.create(
            user=sender_balance.user,
            new_active_balance=sender_balance.active_balance,
            new_frozen_balance=sender_balance.frozen_balance,
            change_active_balance_amount=trader_commission_amount,
            change_reason=BalanceHistory.ChangeReason.PAY_COMMISSION
        )
        net_amount -= trader_commission_amount

    sender_balance.frozen_balance -= transaction.amount
    sender_balance.save()
    BalanceHistory.objects.create(
        user=sender_balance.user,
        new_active_balance=sender_balance.active_balance,
        new_frozen_balance=sender_balance.frozen_balance,
        change_frozen_balance_amount=-transaction.amount,
        change_reason=BalanceHistory.ChangeReason.TAKE_AWAY_FOR_TRANSACTION
    )
    receiver_user = transaction.merchant if transaction.type == transaction.Type.DEPOSIT else transaction.trader
    receiver_balance = Balance.objects.select_for_update().get(user=receiver_user)
    receiver_balance.active_balance += net_amount
    receiver_balance.save()
    BalanceHistory.objects.create(
        user=receiver_balance.user,
        new_active_balance=receiver_balance.active_balance,
        new_frozen_balance=receiver_balance.frozen_balance,
        change_active_balance_amount=net_amount,
        change_reason=BalanceHistory.ChangeReason.GIVE_AWAY_FOR_TRANSACTION
    )


def release_user_balance_for_transaction(user_balance: Balance, transaction: Transaction):
    """
    Release the stated transaction amount from the user's frozen balance to the active one
    if the transaction ends unsuccessfully or the is redirected to the other user.
    """
    user_balance.active_balance += transaction.amount
    user_balance.frozen_balance -= transaction.amount
    user_balance.save()
    BalanceHistory.objects.create(
        user=user_balance.user,
        new_active_balance=user_balance.active_balance,
        new_frozen_balance=user_balance.frozen_balance,
        change_active_balance_amount=transaction.amount,
        change_frozen_balance_amount=-transaction.amount,
        change_reason=BalanceHistory.ChangeReason.RELEASE_AFTER_TRANSACTION
    )


def update_balance_on_refunded_transaction(user_balance: Balance, transaction: Transaction):
    """
    Remove the amount to be refund from the user's frozen balance when the transaction is refunded.
    """
    refunded_amount = transaction.actual_amount - transaction.amount
    user_balance.frozen_balance -= refunded_amount
    user_balance.save()
    BalanceHistory.objects.create(
        user=user_balance.user,
        new_active_balance=user_balance.active_balance,
        new_frozen_balance=user_balance.frozen_balance,
        change_frozen_balance_amount=-refunded_amount,
        change_reason=BalanceHistory.ChangeReason.RELEASE_AFTER_REFUND
    )
    bank_details = transaction.trader_bank_details
    bank_details.current_daily_turnover += refunded_amount
    bank_details.current_weekly_turnover += refunded_amount
    bank_details.current_monthly_turnover += refunded_amount
    bank_details.save()  # TODO: consider this logic for withdrawal transactions (count the commissions)


def update_balances_on_redirect_transaction(sender_balance: Balance, transaction: Transaction):
    """
    Check if there is an old trader assigned to the transaction,
    then release his balance and freeze for the new one.

    - Raise an exception if there is no old trader.
    """
    if hasattr(transaction, 'old_trader'):
        old_sender_balance = Balance.objects.select_for_update().get(user=transaction.old_trader)
        release_user_balance_for_transaction(old_sender_balance, transaction)
        transaction.old_trader = None
        freeze_user_balance_for_transaction(sender_balance, transaction)

        PrevTransactionTraders.objects.create(trader=transaction.old_trader, transaction=transaction)
    else:
        raise ValueError(
            f'Old trader not found to redirect'
            f'transaction ID: {transaction.id} to: {transaction.trader.id}'
        )


@database_sync_to_async
def update_last_seen(user: User):
    # TODO: implement logic to notify admins with trader Online status
    #  and start taking into account it while searching trader for transaction
    with atomic():
        user.last_seen = timezone.now()
        user.save()


@database_sync_to_async
def handle_transaction(user: User, data):
    if not user.is_authenticated:
        raise Exception("User is not authenticated")
    try:
        transaction = Transaction.objects.get(pk=data.get("transaction_id"))
    except Transaction.DoesNotExist:
        raise Exception(f"Transaction ID: {data.get('transaction_id')} does not exist")

    if transaction.trader != user:
        raise Exception("Trader has no permissions to modify this transaction.")

    if (transaction.status == Transaction.Status.REVIEWING
            and transaction.type == Transaction.Type.DEPOSIT or
            transaction.status == Transaction.Status.PENDING and
            transaction.type == Transaction.Type.WITHDRAWAL):
        new_status = data.get('new_status')

        if transaction.type == Transaction.Type.DEPOSIT:
            if new_status in (Transaction.Status.FAILED, Transaction.Status.COMPLETED):
                transaction.status = new_status
            elif new_status in (Transaction.Status.REFUND_REQUESTED, Transaction.Status.PARTIAL):
                actual_amount = data.get('actual_amount')
                if actual_amount is None:
                    raise Exception("Actual amount was not provided to update status")
                transaction.status = new_status
                transaction.actual_amount = actual_amount
            else:
                raise Exception("Trader has no permissions to set the transaction to this status")
        elif transaction.type == Transaction.Type.WITHDRAWAL:
            if new_status in (Transaction.Status.CANCELLED, Transaction.Status.DISPUTING):
                transaction.status = new_status
            else:
                raise Exception("Trader has no permissions to set the transactions to this status")
        transaction.save()
    else:
        raise Exception("Trader has no permissions to change status at this point.")


@database_sync_to_async
def get_current_balance(user: User):
    user = user if user.type != User.Type.ADMIN else None
    balance = Balance.objects.get(user=user)
    return BalanceSerializer(balance).data


@database_sync_to_async
def get_active_transactions(trader: User):
    transactions = Transaction.objects.filter(
        trader=trader,
        status__in=[Transaction.Status.PENDING, Transaction.Status.REVIEWING]
    )
    return TransactionSerializer(transactions, many=True).data


@database_sync_to_async
def settle_transaction(data):
    try:
        transaction = Transaction.objects.get(pk=data.get("transaction_id"))
    except Transaction.DoesNotExist:
        raise Exception(f"Transaction ID: {data.get('transaction_id')} does not exist")

    if transaction.status == Transaction.Status.DISPUTING:
        new_status = data.get("new_status")
        if (new_status == Transaction.Status.COMPLETED or
                new_status == Transaction.Status.PARTIAL or
                new_status == Transaction.Status.FAILED):
            actual_amount = data.get("actual_amount")
            if actual_amount is None:
                raise Exception("Actual amount was not provided to update status")

            transaction.status = new_status
            transaction.actual_amount = actual_amount
            transaction.save()
        else:
            raise Exception("Admin has no permissions to set transactions to this status")
    else:
        raise Exception("Admin has no permissions to change status at this point")


@database_sync_to_async
def client_handle_transaction(user: User, data):
    try:
        transaction = Transaction.objects.get(order_id=data.pop('order_id'), merchant=user)
    except Transaction.DoesNotExist:
        raise Exception('Invalid order_id')

    serializer = ClientTransactionStatusUpdateSerializer(transaction, data=data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return serializer.data


def process_transactions_from_df(merchant_id: int, transactions_df: DataFrame):
    for index, transaction_data in transactions_df.iterrows():
        transaction_dict = transaction_data.to_dict()
        process_transaction_task.delay(transaction_dict, merchant_id)
