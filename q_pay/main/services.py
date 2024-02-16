from typing import List

import openpyxl
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.db.models import F

from .models import User, Transaction, Balance, ServiceBalance, PrevTransactionTraders
from .serializers import TransactionSerializer


def create_transactions_excel(transactions: List[Transaction], transaction_type='input'):
    # TODO: modify in according to the current transaction model
    """
    Prepare an exel file with transactions report.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Transactions'

    if transaction_type == 'input':
        columns = ['Transaction ID', 'Status', 'Merchant', 'Bank', 'Requisites', 'Claimed Amount', 'Actual Amount',
                   'Date']
    elif transaction_type == 'output':
        columns = ['Transaction ID', 'Status', 'Amount', 'Bank', 'Card Number', 'Receipt URL', 'Date']
    else:
        raise ValueError("Invalid transaction type. Must be 'input' or 'output'.")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title

    for row_num, transaction in enumerate(transactions, 2):
        worksheet.cell(row=row_num, column=1).value = str(transaction.id)
        worksheet.cell(row=row_num, column=2).value = transaction.get_status_display()

        if transaction_type == 'input':
            worksheet.cell(row=row_num, column=3).value = transaction.merchant.id
            worksheet.cell(row=row_num, column=4).value = transaction.trader_bank_details.bank.title
            worksheet.cell(row=row_num, column=5).value = transaction.trader_bank_details.card_number
            worksheet.cell(row=row_num, column=6).value = transaction.amount
            worksheet.cell(row=row_num, column=7).value = transaction.actual_amount or 0
        elif transaction_type == 'output':
            worksheet.cell(row=row_num, column=3).value = transaction.amount
            if transaction.trader_bank_details:
                value = transaction.trader_bank_details.bank.title
            else:
                value = 'N/A'
            worksheet.cell(row=row_num, column=4).value = value
            worksheet.cell(row=row_num, column=5).value = transaction.client_card_number

        worksheet.cell(row=row_num, column=8).value = transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')

    return workbook


def get_eligible_trader_ids_for_transactions(transactions: List[Transaction]) -> List[str]:
    # TODO: modify in according to the changes
    """
    Get traders eligible to handle the transactions.

    Search filters:
        1. The trader is active
        2. The trader has active ads with the banks of the transactions
        3. The every single ad has active bank details in itself with eligible unused limits for the transactions
        # TODO: implement
    """
    eligible_trader_ids = User.objects.filter(
        type=User.Type.TRADER,
        is_active=True
    ).values_list('id', flat=True)

    for transaction in transactions:
        eligible_trader_ids = eligible_trader_ids.filter(
            advertisements__is_active=True,
            advertisements__bank_id=transaction.trader_bank_details.bank,
        ).values_list('id', flat=True)

        if not eligible_trader_ids:
            break

    return list(eligible_trader_ids)


def notify_trader_about_new_transaction(transaction_id: str):
    """
    Send an event about a new transaction to notify the trader and provide the transaction data.
    """
    transaction = Transaction.objects.get(pk=transaction_id)
    transaction_data = TransactionSerializer(transaction).data

    channel_layer = get_channel_layer()
    group_name = f'user_{transaction.trader.id}'
    message = {
        'type': 'send_transaction_to_user',
        'transaction_data': transaction_data,
    }
    async_to_sync(channel_layer.group_send)(group_name, message)  # TODO: check fixes to avoid warning


def notify_admins_about_new_transaction(transaction_id: str):
    """
    Send an event about a new transaction to notify all the admins and provide the transaction data.

    - It is intended for use on transaction dispute only.
    """
    transaction = Transaction.objects.get(pk=transaction_id)
    transaction_data = TransactionSerializer(transaction).data

    for admin in User.objects.admins():
        channel_layer = get_channel_layer()
        group_name = f'user_{admin.id}'
        message = {
            'type': 'send_transaction_to_user',
            'transaction_data': transaction_data,
        }
        async_to_sync(channel_layer.group_send)(group_name, message)  # TODO: check fixes to avoid warning


def freeze_user_balance_for_transaction(user_balance: Balance, transaction: Transaction):
    """
    Freeze the stated transaction amount on the user's balance until the end of the transaction.
    """
    user_balance.active_balance = F('active_balance') - transaction.amount
    user_balance.frozen_balance = F('frozen_balance') + transaction.amount


def update_balances_on_successful_transaction(sender_balance: Balance, transaction: Transaction):
    """
    Perform balance transfers between the users if the transaction ends successfully.

    Return the penalty amount to the sender's balance if the transaction actual amount is less than stated amount.

    Freeze the amount to refund on the sender's balance if the transaction actual amount exceeds the stated amount.

    Calculate the amount to be transferred equal the actual amount
        if it does not exceed the stated amount,
        otherwise of the stated amount,
        without taking into account the commission amounts.

    Pay to the service and to the trader the amounts of the
        commissions of the amount to be transferred.

    Calculate the net amount according to the transaction type:
        1. Subtract the service commission amount from the amount to be transferred.
        2. Subtract the trader commission amount as well, if the trader is sender (for deposit transactions).

    Remove the stated amount from the sender's balance and transfer the net amount to the receiver's balance.
    """
    if transaction.status == Transaction.Status.PARTIAL:
        penalty_amount = transaction.amount - transaction.actual_amount
        sender_balance.active_balance = F('active_balance') + penalty_amount
    elif transaction.status == Transaction.Status.REFUND_REQUESTED:
        amount_to_refund = transaction.actual_amount - transaction.amount
        sender_balance.frozen_balance = F('frozen_balance') + amount_to_refund

    amount_to_transfer = min(transaction.actual_amount, transaction.amount)

    service_commission_amount = transaction.service_commission * amount_to_transfer
    print("WE AT SERVICE BALANCE")
    service_balance = ServiceBalance.get_singleton()
    service_balance.balance = F('balance') + service_commission_amount
    service_balance.save(update_fields=['balance'])

    net_amount = amount_to_transfer - service_commission_amount
    if transaction.type == Transaction.Type.DEPOSIT:
        trader_commission_amount = transaction.trader_commission * amount_to_transfer
        sender_balance.active_balance = F('active_balance') + trader_commission_amount
        net_amount -= trader_commission_amount

    sender_balance.frozen_balance = F('frozen_balance') - transaction.amount
    receiver_user = transaction.merchant if transaction.type == transaction.Type.DEPOSIT else transaction.trader
    receiver_balance = Balance.objects.select_for_update().get(user=receiver_user)
    receiver_balance.active_balance = F('active_balance') + net_amount
    receiver_balance.save(update_fields=['active_balance'])


def release_user_balance_for_transaction(user_balance: Balance, transaction: Transaction):
    """
    Release the stated transaction amount from the user's frozen balance to the active one
    if the transaction ends unsuccessfully or the is redirected to the other user.
    """
    user_balance.active_balance = F('active_balance') + transaction.amount
    user_balance.frozen_balance = F('frozen_balance') - transaction.amount


def update_balance_on_refunded_transaction(user_balance: Balance, transaction: Transaction):
    """
    Remove the amount to be refund from the user's frozen balance when the transaction is refunded.
    """
    refunded_amount = transaction.actual_amount - transaction.amount
    user_balance.frozen_balance = F('frozen_balance') - refunded_amount


def update_balances_on_redirect_transaction(sender_balance: Balance, transaction: Transaction):
    """
    Check if there is an old trader assigned to the transaction,
    then release his balance and freeze for the new one.

    - Raise an exception if there is no old trader.
    """
    if hasattr(transaction, 'old_trader'):
        old_sender_balance = Balance.objects.select_for_update().get(user=transaction.old_trader)
        release_user_balance_for_transaction(old_sender_balance, transaction)
        old_sender_balance.save(update_fields=['active_balance', 'frozen_balance'])
        freeze_user_balance_for_transaction(sender_balance, transaction)

        PrevTransactionTraders.objects.create(trader=transaction.old_trader, transaction=transaction)
    else:
        raise ValueError(
            f'Old trader not found to redirect'
            f'transaction ID: {transaction.id} to: {transaction.trader.id}'
        )
